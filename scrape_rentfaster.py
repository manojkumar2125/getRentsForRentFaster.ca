import time
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementClickInterceptedException
from webdriver_manager.chrome import ChromeDriverManager

BASE_URL = "https://www.rentfaster.ca"

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.get("https://www.rentfaster.ca/ab/calgary/rentals/#listview")
wait = WebDriverWait(driver, 20)

# Handle cookie popup
try:
    wait.until(EC.element_to_be_clickable((By.ID, "ez-cookie-notification__accept"))).click()
except TimeoutException:
    pass

time.sleep(3)

# Get total pages
try:
    pagination_info = driver.find_element(By.CSS_SELECTOR, "span.paging-text-span span.ng-binding").text
    total_pages = int(pagination_info.split("of")[-1].strip())
except:
    total_pages = 1

data = []
reference_links = []
seen_links = set()
empty_page_count = 0

for page in range(1, total_pages + 1):
    print(f"\nScraping page {page} ...")

    try:
        listings = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[contains(@class, 'listing-preview-wrap')]")))
    except TimeoutException:
        listings = []

    if not listings:
        empty_page_count += 1
        print(f"⚠️ No listings found on page {page}. Empty count: {empty_page_count}/3")
        if empty_page_count >= 3:
            print("🚨 3 consecutive empty pages. Ending scraping early.")
            break
        continue
    else:
        empty_page_count = 0

    print(f"✅ {len(listings)} listings found.")

    for listing in listings:
        try:
            link_elem = listing.find_element(By.XPATH, ".//a[contains(@class, 'infobox-price')]")
            href = link_elem.get_attribute("href")
            full_link = BASE_URL + href if href.startswith("/") else href
        except:
            full_link = "N/A"

        if full_link in seen_links:
            continue
        seen_links.add(full_link)

        try:
            price_spans = link_elem.find_elements(By.TAG_NAME, "span")
            price_parts = [span.text.strip() for span in price_spans if span.text.strip() and span.text.strip() != "$"]
            price = " ".join(price_parts).replace(" - ", "-")
        except:
            price = "N/A"

        try:
            address = listing.find_element(By.CSS_SELECTOR, "div.is-size-7.dnt.ng-binding.ng-scope").text.strip()
        except:
            address = "N/A"

        try:
            beds = listing.find_element(By.XPATH, ".//div[contains(@class, 'level-item') and contains(., 'bd')]").text.replace("bd", "").strip()
        except:
            beds = "N/A"

        try:
            baths = listing.find_element(By.XPATH, ".//div[contains(@class, 'level-item') and contains(., 'ba')]").text.replace("ba", "").strip()
        except:
            baths = "N/A"

        try:
            area = listing.find_element(By.XPATH, ".//div[contains(@class,'level-item') and contains(., 'ft')]").text.strip()
        except:
            area = "N/A"

        try:
            pets = listing.find_element(By.XPATH, ".//div[contains(@class,'level-item') and contains(text(), 'Pets')]").text.strip()
        except:
            pets = "N/A"

        try:
            house_type = listing.find_element(By.XPATH, ".//div[contains(@class,'is-size-7 ng-binding')][.//i[contains(@class, 'fa-home')]]").text.strip()
        except:
            house_type = "N/A"

        try:
            community = listing.find_element(By.XPATH, ".//div[contains(@class,'is-size-7 dnt ng-binding')][.//i[contains(@class, 'fa-users')]]").text.strip()
        except:
            community = "N/A"

        data.append({
            "Price": price,
            "Address": address,
            "Beds": beds,
            "Baths": baths,
            "Area": area,
            "Pets Allowed": pets,
            "House Type": house_type,
            "Community": community,
            "Reference Link": full_link
        })

        reference_links.append(full_link)

    # Move to next page
    if page < total_pages:
        next_page_number = page + 1
        try:
            xpath = f"//a[contains(@class, 'button') and text()='{next_page_number}']"
            next_page_btn = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
            driver.execute_script("arguments[0].scrollIntoView(true);", next_page_btn)
            next_page_btn.click()
            time.sleep(4)
        except (TimeoutException, NoSuchElementException, ElementClickInterceptedException) as e:
            print(f"Could not go to page {next_page_number}: {e}")
            break

driver.quit()

# Save to Excel
wb = Workbook()
ws1 = wb.active
ws1.title = "Listings"

df_main = pd.DataFrame(data)
for r in dataframe_to_rows(df_main, index=False, header=True):
    ws1.append(r)

ws2 = wb.create_sheet(title="Reference")
ws2.append(["Reference Link"])
for link in reference_links:
    cell = ws2.cell(row=ws2.max_row + 1, column=1, value=link)
    cell.hyperlink = link
    cell.font = Font(color="0000FF", underline="single")

filename = "calgary_rentals.xlsx"
wb.save(filename)

print(f"\n📦 Total listings collected: {len(data)}")
print(f"📁 Saved to file: {filename}")
